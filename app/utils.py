from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build_budget_excel(entries, summary, current_balance, selected_month):
    wb = Workbook()
    ws = wb.active
    ws.title = "Бюджет"

    title_fill = PatternFill("solid", fgColor="2563EB")
    section_fill = PatternFill("solid", fgColor="EEF2FF")
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    income_fill = PatternFill("solid", fgColor="ECFDF3")
    expense_fill = PatternFill("solid", fgColor="FEF2F2")

    white_font = Font(color="FFFFFF", bold=True, size=14)
    title_font = Font(bold=True, size=16)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовок
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Бюджет за месяц: {selected_month}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill
    ws["A1"].font = white_font

    # Сводка
    ws.merge_cells("A3:D3")
    ws["A3"] = "Сводка"
    ws["A3"].font = bold_font
    ws["A3"].fill = section_fill
    ws["A3"].alignment = left

    summary_rows = [
        ("Доход за месяц", summary["income"]),
        ("Расход за месяц", summary["expense"]),
        ("Баланс за месяц", summary["balance"]),
        ("Текущий баланс", current_balance),
    ]

    row = 4
    for label, value in summary_rows:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = int(round(value))
        ws[f"A{row}"].font = bold_font
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border
        ws[f"B{row}"].number_format = '# ##0 "Р"'
        row += 1

    # Таблица операций
    table_start = row + 2
    ws.merge_cells(f"A{table_start}:D{table_start}")
    ws[f"A{table_start}"] = "Операции"
    ws[f"A{table_start}"].font = bold_font
    ws[f"A{table_start}"].fill = section_fill
    ws[f"A{table_start}"].alignment = left

    headers_row = table_start + 1
    headers = ["Месяц", "Тип", "Категория", "Сумма"]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=headers_row, column=col_index)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    data_row = headers_row + 1
    for entry in entries:
        ws.cell(row=data_row, column=1, value=entry["month_name"])
        ws.cell(row=data_row, column=2, value=entry["entry_type"])
        ws.cell(row=data_row, column=3, value=entry["category"])

        amount_cell = ws.cell(row=data_row, column=4, value=int(round(entry["amount"])))
        amount_cell.number_format = '# ##0 "Р"'

        for col in range(1, 5):
            cell = ws.cell(row=data_row, column=col)
            cell.border = border
            cell.alignment = left

        if entry["entry_type"] == "Доход":
            for col in range(1, 5):
                ws.cell(row=data_row, column=col).fill = income_fill
        else:
            for col in range(1, 5):
                ws.cell(row=data_row, column=col).fill = expense_fill

        data_row += 1

    # Ширина колонок
    widths = {
        1: 18,
        2: 16,
        3: 28,
        4: 16,
    }

    for col_index, width in widths.items():
        ws.column_dimensions[get_column_letter(col_index)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_shootings_excel(shootings, report_title):
    wb = Workbook()
    ws = wb.active
    ws.title = "Съёмки"

    title_fill = PatternFill("solid", fgColor="14532D")
    section_fill = PatternFill("solid", fgColor="DCFCE7")
    header_fill = PatternFill("solid", fgColor="BBF7D0")
    odd_row_fill = PatternFill("solid", fgColor="F0FDF4")
    even_row_fill = PatternFill("solid", fgColor="FFFFFF")

    white_bold_font = Font(color="FFFFFF", bold=True, size=14)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:J1")
    ws["A1"] = report_title
    ws["A1"].fill = title_fill
    ws["A1"].font = white_bold_font
    ws["A1"].alignment = center

    ws.merge_cells("A3:J3")
    ws["A3"] = f"Всего записей: {len(shootings)}"
    ws["A3"].fill = section_fill
    ws["A3"].font = bold_font
    ws["A3"].alignment = left

    headers = [
        "Дата",
        "Время",
        "Название",
        "Клиент",
        "Телефон",
        "Часы",
        "Стоимость",
        "Предоплата",
        "Остаток",
        "Комментарий",
    ]

    headers_row = 5
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=headers_row, column=col_index)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    data_row = headers_row + 1
    for index, shooting in enumerate(shootings):
        price = int(round(float(shooting.get("price") or 0)))
        prepayment = int(round(float(shooting.get("prepayment") or 0)))
        remaining_payment = max(price - prepayment, 0)

        row_fill = odd_row_fill if index % 2 == 0 else even_row_fill

        values = [
            shooting.get("shooting_date_display") or shooting.get("shooting_date") or "—",
            shooting.get("shooting_time") or "—",
            shooting.get("project_name") or "—",
            shooting.get("client_name") or "—",
            shooting.get("phone") or "—",
            shooting.get("duration_hours") or "—",
            price,
            prepayment,
            remaining_payment,
            shooting.get("notes") or "—",
        ]

        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_index, value=value)
            cell.border = border
            cell.alignment = left if col_index != 6 else center
            cell.fill = row_fill

            if col_index in {7, 8, 9}:
                cell.number_format = '# ##0 "Р"'

        data_row += 1

    column_widths = {
        1: 18,
        2: 10,
        3: 24,
        4: 22,
        5: 18,
        6: 10,
        7: 14,
        8: 14,
        9: 14,
        10: 36,
    }
    for col_index, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_index)].width = width

    ws.freeze_panes = "A6"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_photo_project_excel(project, bookings):
    wb = Workbook()
    ws = wb.active
    ws.title = "Фотопроект"

    title_fill = PatternFill("solid", fgColor="2563EB")
    section_fill = PatternFill("solid", fgColor="EEF2FF")
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    odd_row_fill = PatternFill("solid", fgColor="F8FAFF")
    even_row_fill = PatternFill("solid", fgColor="FFFFFF")

    white_bold_font = Font(color="FFFFFF", bold=True, size=14)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:K1")
    ws["A1"] = f"Фотопроект: {project.get('title') or 'Без названия'}"
    ws["A1"].fill = title_fill
    ws["A1"].font = white_bold_font
    ws["A1"].alignment = center

    ws.merge_cells("A3:K3")
    ws["A3"] = (
        f"Город: {project.get('city') or '—'} | "
        f"Дата: {project.get('project_date_display') or project.get('project_date') or '—'} | "
        f"Время: {project.get('time_range_display') or '—'}"
    )
    ws["A3"].fill = section_fill
    ws["A3"].font = bold_font
    ws["A3"].alignment = left

    ws.merge_cells("A4:K4")
    ws["A4"] = f"Адрес: {project.get('address') or '—'}"
    ws["A4"].fill = section_fill
    ws["A4"].font = bold_font
    ws["A4"].alignment = left

    ws.merge_cells("A5:K5")
    ws["A5"] = f"Всего записей: {len(bookings)}"
    ws["A5"].fill = section_fill
    ws["A5"].font = bold_font
    ws["A5"].alignment = left

    headers = [
        "Клиент",
        "Контакты",
        "Дата",
        "Время",
        "Длительность",
        "Начало макияжа",
        "Стоимость",
        "Предоплата",
        "Остаток",
        "Статус",
        "Комментарий",
    ]

    headers_row = 7
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=headers_row, column=col_index)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    data_row = headers_row + 1
    for index, booking in enumerate(bookings):
        price = int(round(float(booking.get("price") or 0)))
        prepayment = int(round(float(booking.get("prepayment") or 0)))
        remaining_payment = max(price - prepayment, 0)
        row_fill = odd_row_fill if index % 2 == 0 else even_row_fill

        values = [
            booking.get("client_name") or "—",
            booking.get("client_contact") or "—",
            booking.get("booking_date_display") or booking.get("booking_date") or "—",
            booking.get("booking_time") or "—",
            f"{booking.get('duration_minutes') or 15} мин",
            booking.get("makeup_start_time") or "—",
            price,
            prepayment,
            remaining_payment,
            booking.get("status") or "—",
            booking.get("comment") or "—",
        ]

        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_index, value=value)
            cell.border = border
            cell.alignment = center if col_index in {4, 5, 6, 10} else left
            cell.fill = row_fill
            if col_index in {7, 8, 9}:
                cell.number_format = '# ##0 "Р"'
        data_row += 1

    column_widths = {
        1: 22,
        2: 24,
        3: 16,
        4: 10,
        5: 14,
        6: 16,
        7: 14,
        8: 14,
        9: 14,
        10: 14,
        11: 28,
    }
    for col_index, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_index)].width = width

    ws.freeze_panes = "A8"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
