import base64
from datetime import datetime, date
from io import BytesIO
import hmac
import os
import re
import secrets

from flask import (
    current_app,
    flash,
    redirect,
    render_template,
    request,
    session,
    send_file,
    url_for,
)
from flask_login import (
    current_user,
    login_required,
    login_user,
    logout_user,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

from app.auth import (
    generate_totp_secret,
    load_user_from_db,
    verify_user,
    verify_totp_code,
    is_login_rate_limited,
    register_failed_login,
    clear_failed_logins,
)
from config import Config
from app.models import (
    archive_car_notification,
    create_budget_entry,
    create_car_done_service,
    create_car_planned_service,
    create_car_planned_service_from_notification,
    create_shooting,
    delete_archived_car_notification,
    delete_budget_entry,
    delete_car_done_service,
    delete_car_planned_service,
    delete_shooting,
    get_all_budget_entries,
    get_archived_car_notifications,
    get_balance_for_month,
    get_balance_history,
    get_budget_entry_by_id,
    get_budget_summary,
    get_car_done_service_by_id,
    get_car_done_services,
    get_car_last_mileage,
    get_car_planned_service_by_id,
    get_car_planned_services,
    get_car_total_spent,
    get_current_balance,
    get_hidden_notification_keys,
    get_periodic_services_for_notifications,
    get_shooting_by_id,
    get_upcoming_shootings,
    hide_car_notification,
    is_car_notification_hidden,
    move_planned_to_done,
    append_car_services,
    replace_budget_entries,
    replace_car_services,
    replace_shootings,
    save_balance_history,
    set_current_balance,
    update_budget_entry,
    update_car_done_service,
    update_car_planned_service,
    update_shooting,
    create_shooting,
    delete_shooting,
    get_shooting_by_id,
    get_upcoming_shootings,
    update_shooting,
    create_shooting,
    delete_shooting,
    get_archived_shootings,
    get_nearest_shooting,
    get_shooting_by_id,
    get_shootings_count,
    get_upcoming_shootings,
    create_scenario,
    delete_scenario,
    get_archived_scenarios,
    get_nearest_scenario,
    get_scenario_by_id,
    get_scenarios_count,
    get_upcoming_scenarios,
    toggle_scenario_status,
    update_scenario,
    update_shooting,
    get_user_by_id,
    set_user_avatar_filename,
    set_user_last_login_ip,
    set_user_otp,
    set_user_password_hash,
)
from app.utils import build_budget_excel, build_shootings_excel, build_scenarios_excel
from app.logging_setup import log_audit, log_invalid_form, log_import_result
from app.planner import (
    create_task,
    delete_task_for_scenario,
    delete_task_for_shooting,
    replace_manual_schedule_tasks,
    upsert_task_for_scenario,
    upsert_task_for_shooting,
)


MONTHS = [
    "Январь",
    "Февраль",
    "Март",
    "Апрель",
    "Май",
    "Июнь",
    "Июль",
    "Август",
    "Сентябрь",
    "Октябрь",
    "Ноябрь",
    "Декабрь",
]

BASE_CATEGORIES = ["Авто", "Еда", "Другое"]

MONTH_MAP = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}


def _get_current_month_name() -> str:
    return MONTH_MAP[datetime.now().month]


def _get_period_months(period_type: str) -> int:
    if period_type == "6 мес":
        return 6
    if period_type == "12 мес":
        return 12
    return 0


def _add_months(base_date: datetime, months: int) -> datetime:
    month_index = base_date.month - 1 + months
    year = base_date.year + month_index // 12
    month = month_index % 12 + 1
    return datetime(year, month, 1)


def _parse_selected_ids(raw_value: str) -> list[int]:
    if not raw_value:
        return []
    result = []
    for item in raw_value.split(","):
        token = item.strip()
        if token.isdigit():
            result.append(int(token))
    return result


def _build_car_notifications():
    periodic_planned, periodic_done = get_periodic_services_for_notifications()
    hidden_keys = set(get_hidden_notification_keys())
    archived_rows = get_archived_car_notifications()
    notifications = []
    today = datetime.now().replace(day=1)

    month_names = {
        "01": "январь",
        "02": "февраль",
        "03": "март",
        "04": "апрель",
        "05": "май",
        "06": "июнь",
        "07": "июль",
        "08": "август",
        "09": "сентябрь",
        "10": "октябрь",
        "11": "ноябрь",
        "12": "декабрь",
    }

    for item in periodic_planned:
        notification_key = f"planned:{item['id']}"
        if notification_key in hidden_keys:
            continue

        notifications.append({
            "notification_key": notification_key,
            "title": item["service_name"],
            "status": "Скоро",
            "period_type": item["period_type"],
            "detail_description": item["detail_description"] or "—",
            "last_service_date_text": "—",
            "work_kind": item["work_kind"] if "work_kind" in item.keys() and item["work_kind"] else "",
        })

    for item in periodic_done:
        period_months = _get_period_months(item["period_type"])
        if not period_months or not item["service_date"]:
            continue

        try:
            last_date = datetime.strptime(item["service_date"], "%Y-%m")
        except ValueError:
            continue

        next_service_date = _add_months(last_date, period_months)
        status = "Нужна замена" if today >= next_service_date else "Скоро"
        last_service_date_text = f"{month_names[last_date.strftime('%m')]} {last_date.year}"

        notification_key = f"done:{item['id']}:{item['service_date']}"
        if notification_key in hidden_keys:
            continue

        notifications.append({
            "notification_key": notification_key,
            "title": item["service_name"],
            "status": status,
            "period_type": item["period_type"],
            "detail_description": item["detail_description"] or "—",
            "last_service_date_text": last_service_date_text,
            "work_kind": item["work_kind"] if "work_kind" in item.keys() and item["work_kind"] else "",
        })

    for item in archived_rows:
        notifications.append({
            "notification_key": item["notification_key"],
            "title": item["title"],
            "status": "Архив",
            "period_type": item["period_type"] or "—",
            "detail_description": item["detail_description"] or "—",
            "last_service_date_text": item["last_service_date_text"] or "—",
            "work_kind": item["work_kind"] or "",
        })

    status_priority = {
        "Нужна замена": 0,
        "Скоро": 1,
        "Архив": 2,
    }

    notifications.sort(
        key=lambda item: (
            status_priority.get(item["status"], 9),
            item["title"].lower(),
        )
    )

    return notifications





def _format_shooting_date(date_value: str) -> str:
    if not date_value:
        return "—"

    try:
        parsed = datetime.strptime(date_value, "%Y-%m-%d")
        month_names = {
            1: "января",
            2: "февраля",
            3: "марта",
            4: "апреля",
            5: "мая",
            6: "июня",
            7: "июля",
            8: "августа",
            9: "сентября",
            10: "октября",
            11: "ноября",
            12: "декабря",
        }
        return f"{parsed.day} {month_names[parsed.month]} {parsed.year}"
    except ValueError:
        return date_value


def _format_shooting_time(time_value: str) -> str:
    return time_value if time_value else "Без времени"


def _is_mobile_request() -> bool:
    user_agent = request.headers.get("User-Agent", "").lower()
    mobile_markers = (
        "android",
        "iphone",
        "ipad",
        "ipod",
        "windows phone",
        "opera mini",
        "mobile",
    )
    return any(marker in user_agent for marker in mobile_markers)


def _prepare_shooting_row(row):
    shooting = dict(row)
    shooting["shooting_date_display"] = _format_shooting_date(shooting.get("shooting_date", ""))
    shooting["shooting_time_display"] = _format_shooting_time(shooting.get("shooting_time", ""))

    try:
        price_value = float(shooting.get("price") or 0)
    except (TypeError, ValueError):
        price_value = 0

    try:
        prepayment_value = float(shooting.get("prepayment") or 0)
    except (TypeError, ValueError):
        prepayment_value = 0

    shooting["price"] = price_value
    shooting["prepayment"] = prepayment_value
    shooting["remaining_payment"] = max(price_value - prepayment_value, 0)

    return shooting


def _get_shooting_form_data(form):
    return {
        "client_name": form.get("client_name", "").strip(),
        "project_name": form.get("project_name", "").strip(),
        "shooting_date": form.get("shooting_date", "").strip(),
        "shooting_time": form.get("shooting_time", "").strip(),
        "location": form.get("location", "").strip(),
        "package_name": form.get("package_name", "").strip(),
        "status": form.get("status", "Запланирована").strip() or "Запланирована",
        "phone": form.get("phone", "").strip(),
        "price": form.get("price", "0").strip(),
        "prepayment": form.get("prepayment", "0").strip(),
        "notes": form.get("notes", "").strip(),
    }


def _validate_shooting_form(data):
    errors = []

    if not data["client_name"]:
        errors.append("Укажи имя клиента.")

    if not data["shooting_date"]:
        errors.append("Укажи дату съёмки.")
    else:
        try:
            datetime.strptime(data["shooting_date"], "%Y-%m-%d")
        except ValueError:
            errors.append("Дата съёмки указана в неверном формате.")

    if data["shooting_time"]:
        try:
            datetime.strptime(data["shooting_time"], "%H:%M")
        except ValueError:
            errors.append("Время съёмки указано в неверном формате.")

    if data["status"] not in SHOOTING_STATUSES:
        errors.append("Выбран неизвестный статус съёмки.")

    for field_name, field_label in (("price", "Стоимость"), ("prepayment", "Предоплата")):
        try:
            float(data[field_name] or 0)
        except ValueError:
            errors.append(f"Поле '{field_label}' должно быть числом.")

    return errors


def _is_valid_phone_number(phone_value: str) -> bool:
    if not phone_value:
        return True
    normalized = re.sub(r"[\s\-\(\)]", "", phone_value)
    return bool(re.fullmatch(r"\+?\d{7,15}", normalized))


def _is_valid_contact_value(contact_value: str) -> bool:
    if not contact_value:
        return True
    if len(contact_value) > 255:
        return False
    if _is_valid_phone_number(contact_value):
        return True
    return bool(
        re.fullmatch(
            r"(https?://\S+|www\.\S+|t\.me/\S+|@\w{3,}|[\w.\-]+@[\w.\-]+\.\w+)",
            contact_value.strip(),
        )
    )


def _parse_non_negative_number(value: str, label: str):
    if value in (None, ""):
        return None, None
    try:
        parsed_value = float(value)
    except ValueError:
        return None, f"Поле '{label}' должно быть числом."

    if parsed_value < 0:
        return None, f"Поле '{label}' не может быть отрицательным."

    return parsed_value, None


def _normalize_excel_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _parse_excel_number(value):
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return value
    try:
        return float(str(value).replace(",", ".").strip())
    except ValueError:
        return 0


def _parse_excel_date(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m", "%Y-%m-%d", "%d.%m.%Y", "%m.%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%Y-%m")
        except ValueError:
            continue
    return text


def _parse_excel_full_date(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m", "%m.%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in ("%Y-%m", "%m.%Y"):
                return parsed.strftime("%Y-%m-01")
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text


def _parse_excel_time(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, date):
        return ""

    text = str(value).strip()
    if not text:
        return ""

    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%H:%M")
        except ValueError:
            continue

    return text


def _parse_car_excel(file_storage):
    workbook = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Файл Excel пустой.")

    headers = [_normalize_excel_header(value) for value in rows[0]]
    header_map = {name: index for index, name in enumerate(headers) if name}

    service_name_index = header_map.get("наименование", header_map.get("service_name"))
    if service_name_index is None:
        raise ValueError("В Excel не найден столбец 'Наименование'.")

    description_index = header_map.get("описание", header_map.get("detail_description"))
    period_index = header_map.get("периодичность", header_map.get("period_type"))
    status_index = header_map.get("статус", header_map.get("status"))
    date_index = header_map.get("дата", header_map.get("service_date"))
    mileage_index = header_map.get("пробег", header_map.get("mileage"))
    cost_index = header_map.get("стоимость", header_map.get("service_cost"))

    done_services = []
    planned_services = []

    for row in rows[1:]:
        service_name = str(row[service_name_index]).strip() if row[service_name_index] is not None else ""
        if not service_name:
            continue

        detail_description = ""
        period_type = ""
        status_text = ""
        service_date = ""
        mileage_value = 0
        cost_value = 0

        if description_index is not None and description_index < len(row):
            detail_description = str(row[description_index]).strip() if row[description_index] is not None else ""
        if period_index is not None and period_index < len(row):
            period_type = str(row[period_index]).strip() if row[period_index] is not None else ""
        if status_index is not None and status_index < len(row):
            status_text = str(row[status_index]).strip() if row[status_index] is not None else ""
        if date_index is not None and date_index < len(row):
            service_date = _parse_excel_date(row[date_index])
        if mileage_index is not None and mileage_index < len(row):
            mileage_value = _parse_excel_number(row[mileage_index])
        if cost_index is not None and cost_index < len(row):
            cost_value = _parse_excel_number(row[cost_index])

        normalized_status = status_text.lower()
        is_done = normalized_status == "выполнено" or bool(service_date)

        if is_done:
            done_services.append({
                "service_name": service_name,
                "service_cost": cost_value,
                "mileage": mileage_value,
                "service_date": service_date,
                "detail_description": detail_description,
                "work_kind": "",
                "period_type": period_type,
                "status": "Выполнено",
            })
        else:
            planned_services.append({
                "service_name": service_name,
                "planned_cost": cost_value,
                "mileage": mileage_value,
                "detail_description": detail_description,
                "work_kind": "",
                "period_type": period_type,
                "status": "Планируется",
            })

    return done_services, planned_services


def _parse_budget_excel(file_storage):
    workbook = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Файл Excel пустой.")

    headers = [_normalize_excel_header(value) for value in rows[0]]
    header_map = {name: index for index, name in enumerate(headers) if name}

    month_index = header_map.get("месяц", header_map.get("month_name"))
    type_index = header_map.get("тип", header_map.get("entry_type"))
    category_index = header_map.get("категория", header_map.get("category"))
    amount_index = header_map.get("сумма", header_map.get("amount"))

    if month_index is None or type_index is None or category_index is None or amount_index is None:
        raise ValueError("Для импорта бюджета нужны столбцы: Месяц, Тип, Категория, Сумма.")

    entries = []
    for row in rows[1:]:
        month_name = str(row[month_index]).strip() if month_index < len(row) and row[month_index] is not None else ""
        entry_type = str(row[type_index]).strip() if type_index < len(row) and row[type_index] is not None else ""
        category = str(row[category_index]).strip() if category_index < len(row) and row[category_index] is not None else ""
        amount_raw = row[amount_index] if amount_index < len(row) else None

        if not month_name and not entry_type and not category and amount_raw in (None, ""):
            continue

        if not month_name or not entry_type or not category:
            raise ValueError("В строках бюджета обязательны поля: Месяц, Тип, Категория.")

        amount_value = _parse_excel_number(amount_raw)
        entries.append({
            "month_name": month_name,
            "entry_type": entry_type,
            "category": category,
            "amount": int(amount_value),
        })

    return entries


def _parse_shootings_excel(file_storage):
    workbook = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Файл Excel пустой.")

    headers = [_normalize_excel_header(value) for value in rows[0]]
    header_map = {name: index for index, name in enumerate(headers) if name}

    project_index = header_map.get("название", header_map.get("project_name"))
    client_index = header_map.get("клиент", header_map.get("client_name"))
    date_index = header_map.get("дата", header_map.get("shooting_date"))
    time_index = header_map.get("время", header_map.get("shooting_time"))
    duration_index = header_map.get("часы", header_map.get("duration_hours"))
    phone_index = header_map.get("телефон", header_map.get("phone"))
    price_index = header_map.get("стоимость", header_map.get("price"))
    prepayment_index = header_map.get("предоплата", header_map.get("prepayment"))
    notes_index = header_map.get("комментарий", header_map.get("notes"))

    if project_index is None or client_index is None or date_index is None:
        raise ValueError("Для импорта съёмок нужны столбцы: Название, Клиент, Дата.")

    shootings = []
    for row in rows[1:]:
        project_name = str(row[project_index]).strip() if project_index < len(row) and row[project_index] is not None else ""
        client_name = str(row[client_index]).strip() if client_index < len(row) and row[client_index] is not None else ""
        shooting_date = _parse_excel_full_date(row[date_index]) if date_index < len(row) else ""

        if not project_name and not client_name and not shooting_date:
            continue
        if not project_name or not client_name or not shooting_date:
            raise ValueError("В строках съёмок обязательны поля: Название, Клиент, Дата.")

        shooting_time = ""
        duration_hours = 1
        phone = ""
        price = 0
        prepayment = 0
        notes = ""

        if time_index is not None and time_index < len(row) and row[time_index] is not None:
            shooting_time = _parse_excel_time(row[time_index])
        if duration_index is not None and duration_index < len(row):
            duration_hours = _parse_excel_number(row[duration_index]) or 1
        if phone_index is not None and phone_index < len(row) and row[phone_index] is not None:
            phone = str(row[phone_index]).strip()
        if price_index is not None and price_index < len(row):
            price = _parse_excel_number(row[price_index]) or 0
        if prepayment_index is not None and prepayment_index < len(row):
            prepayment = _parse_excel_number(row[prepayment_index]) or 0
        if notes_index is not None and notes_index < len(row) and row[notes_index] is not None:
            notes = str(row[notes_index]).strip()

        shootings.append({
            "project_name": project_name,
            "client_name": client_name,
            "shooting_date": shooting_date,
            "shooting_time": shooting_time,
            "duration_hours": duration_hours,
            "phone": phone,
            "price": price,
            "prepayment": prepayment,
            "notes": notes,
        })

    return shootings


def _parse_schedule_excel(file_storage):
    workbook = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Файл Excel пустой.")

    headers = [_normalize_excel_header(value) for value in rows[0]]
    header_map = {name: index for index, name in enumerate(headers) if name}

    title_index = header_map.get("название", header_map.get("title"))
    date_index = header_map.get("дата", header_map.get("task_date"))
    if title_index is None or date_index is None:
        raise ValueError("Для импорта Графика нужны столбцы: Название, Дата.")

    time_index = header_map.get("время", header_map.get("start_time"))
    description_index = header_map.get("описание", header_map.get("description"))
    importance_index = header_map.get("важность", header_map.get("is_important"))
    range_end_index = header_map.get("дата окончания", header_map.get("range_end_date"))
    status_index = header_map.get("статус", header_map.get("status"))
    type_index = header_map.get("тип", header_map.get("task_type"))

    parsed_tasks = []
    for row in rows[1:]:
        title = str(row[title_index]).strip() if title_index < len(row) and row[title_index] is not None else ""
        task_date = _parse_excel_full_date(row[date_index]) if date_index < len(row) else ""

        if not title and not task_date:
            continue
        if not title or not task_date:
            raise ValueError("В строках Графика обязательны поля: Название, Дата.")

        start_time = ""
        description = ""
        is_important = 0
        range_end_date = ""
        status = "planned"
        task_type = "Личное"

        if time_index is not None and time_index < len(row) and row[time_index] is not None:
            start_time = str(row[time_index]).strip()
        if description_index is not None and description_index < len(row) and row[description_index] is not None:
            description = str(row[description_index]).strip()
        if importance_index is not None and importance_index < len(row) and row[importance_index] is not None:
            importance_value = str(row[importance_index]).strip().lower()
            is_important = 1 if importance_value in {"1", "yes", "да", "важно", "true"} else 0
        if range_end_index is not None and range_end_index < len(row) and row[range_end_index] is not None:
            range_end_date = _parse_excel_full_date(row[range_end_index])
        if status_index is not None and status_index < len(row) and row[status_index] is not None:
            status_raw = str(row[status_index]).strip().lower()
            if status_raw in {"done", "выполнено"}:
                status = "done"
            elif status_raw in {"cancelled", "отменено"}:
                status = "cancelled"
            elif status_raw in {"протяженная", "протяжённая"}:
                status = "planned"
            else:
                status = "planned"
        if type_index is not None and type_index < len(row) and row[type_index] is not None:
            type_value = str(row[type_index]).strip()
            if type_value:
                task_type = type_value

        parsed_tasks.append({
            "title": title,
            "task_date": task_date,
            "description": description,
            "start_time": start_time,
            "is_important": is_important,
            "range_end_date": range_end_date,
            "status": status,
            "task_type": task_type,
        })

    return parsed_tasks


def _resolve_budget_category(form):
    """
    Возвращает итоговую категорию бюджета.
    Если выбрано 'Другое', берём custom_category.
    """
    category = form.get("category", "").strip()
    custom_category = form.get("custom_category", "").strip()

    if category == "Другое":
        return custom_category.strip()

    return category.strip()


def _collect_budget_categories(entries):
    """
    Собирает категории для фильтров:
    сначала базовые, затем все реальные категории из записей.
    """
    categories = ["Авто", "Еда"]
    existing = set(categories)

    for entry in entries:
        category = entry["category"].strip()
        if category and category not in existing:
            categories.append(category)
            existing.add(category)

    if "Другое" not in categories:
        categories.append("Другое")

    return categories


def register_routes(app):
    def _build_avatar_letter(username: str) -> str:
        return (username or "?").strip()[:1].upper() or "?"

    def _build_avatar_url(user_row) -> str | None:
        if not user_row:
            return None
        avatar_filename = (user_row.get("avatar_filename") or "").strip()
        if not avatar_filename:
            return None
        return url_for("static", filename=f"uploads/avatars/{avatar_filename}")

    def _detect_device_and_browser(user_agent_string: str | None):
        ua = (user_agent_string or "").lower()

        os_name = "Неизвестная ОС"
        if "android" in ua:
            os_name = "Android"
        elif "iphone" in ua or "ipad" in ua or "ios" in ua:
            os_name = "iOS"
        elif "windows" in ua:
            os_name = "Windows"
        elif "mac os x" in ua or "macintosh" in ua:
            os_name = "macOS"
        elif "linux" in ua:
            os_name = "Linux"

        if any(token in ua for token in ["mobile", "iphone", "android"]):
            device_type = "Смартфон"
        elif "ipad" in ua or "tablet" in ua:
            device_type = "Планшет"
        else:
            device_type = "ПК"

        browser = "Неизвестный браузер"
        browser_patterns = [
            ("Edge", r"edg/([0-9\.]+)"),
            ("Opera", r"opr/([0-9\.]+)"),
            ("Chrome", r"chrome/([0-9\.]+)"),
            ("Firefox", r"firefox/([0-9\.]+)"),
            ("Safari", r"version/([0-9\.]+).*safari"),
        ]
        for browser_name, pattern in browser_patterns:
            match = re.search(pattern, ua)
            if match:
                browser = f"{browser_name} {match.group(1).split('.')[0]}"
                break

        return f"{device_type} · {os_name}", browser

    @app.route("/")
    @login_required
    def index():
        return render_template("index.html", username=current_user.username)

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for("index"))

        pending_user_id = session.get("pending_login_user_id")
        pending_remember_me = bool(session.get("pending_login_remember"))
        pending_user_row = get_user_by_id(pending_user_id) if pending_user_id else None

        if pending_user_id and (not pending_user_row or not pending_user_row.get("otp_enabled")):
            session.pop("pending_login_user_id", None)
            session.pop("pending_login_remember", None)
            pending_user_row = None
            pending_remember_me = False

        if request.method == "POST":
            stage = request.form.get("stage", "credentials")
            ip_address = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()

            if stage == "otp":
                otp_code = request.form.get("otp_code", "").strip()
                if not pending_user_row:
                    flash("Сессия входа истекла. Введите логин и пароль снова.", "warning")
                    return render_template("login.html", login_stage="credentials"), 400

                username = pending_user_row.get("username", "")
                if is_login_rate_limited(username, ip_address):
                    current_app.logger.warning(
                        "Login rate limit triggered: username=%s ip=%s",
                        username,
                        ip_address,
                    )
                    flash("Слишком много попыток входа. Повторите позже.", "danger")
                    return render_template(
                        "login.html",
                        login_stage="otp",
                        pending_username=username,
                    ), 429

                if not verify_totp_code(pending_user_row.get("otp_secret"), otp_code):
                    register_failed_login(username, ip_address)
                    current_app.logger.warning(
                        "Failed login (invalid otp): username=%s ip=%s",
                        username,
                        ip_address,
                    )
                    flash("Неверный код Google Authenticator.", "danger")
                    return render_template(
                        "login.html",
                        login_stage="otp",
                        pending_username=username,
                    ), 401

                authenticated_user = load_user_from_db(pending_user_row["id"])
                if not authenticated_user:
                    session.pop("pending_login_user_id", None)
                    session.pop("pending_login_remember", None)
                    flash("Пользователь не найден.", "danger")
                    return render_template("login.html", login_stage="credentials"), 400

                clear_failed_logins(username, ip_address)
                session.pop("pending_login_user_id", None)
                session.pop("pending_login_remember", None)
                login_user(authenticated_user, remember=pending_remember_me)
                set_user_last_login_ip(authenticated_user.id, ip_address)
                log_audit(current_app, "user_login", username=authenticated_user.username)
                return redirect(url_for("index"))

            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()
            remember_me = request.form.get("remember_me") == "on"

            if is_login_rate_limited(username, ip_address):
                current_app.logger.warning(
                    "Login rate limit triggered: username=%s ip=%s",
                    username,
                    ip_address,
                )
                flash("Слишком много попыток входа. Повторите позже.", "danger")
                return render_template("login.html", login_stage="credentials"), 429

            user = verify_user(username, password)
            if user:
                db_user = get_user_by_id(user.id)
                should_request_otp = False
                if db_user and db_user.get("otp_enabled"):
                    last_login_ip = (db_user.get("last_login_ip") or "").strip()
                    should_request_otp = not last_login_ip or last_login_ip != ip_address

                if should_request_otp:
                    session["pending_login_user_id"] = user.id
                    session["pending_login_remember"] = 1 if remember_me else 0
                    flash("Логин и пароль верны. Введите код Google Authenticator.", "info")
                    return render_template(
                        "login.html",
                        login_stage="otp",
                        pending_username=user.username,
                    )

                clear_failed_logins(username, ip_address)
                session.pop("pending_login_user_id", None)
                session.pop("pending_login_remember", None)
                login_user(user, remember=remember_me)
                set_user_last_login_ip(user.id, ip_address)
                log_audit(current_app, "user_login", username=user.username)
                return redirect(url_for("index"))

            register_failed_login(username, ip_address)
            current_app.logger.warning("Failed login: username=%s ip=%s", username, ip_address)
            flash("Неверный логин или пароль.", "danger")
            return render_template("login.html", login_stage="credentials"), 401

        if pending_user_row and pending_user_row.get("otp_enabled"):
            return render_template(
                "login.html",
                login_stage="otp",
                pending_username=pending_user_row["username"],
            )

        return render_template("login.html", login_stage="credentials")

    @app.route("/setup-2fa", methods=["GET", "POST"])
    def setup_2fa():
        pending_user_id = session.get("pending_otp_user_id")
        from_settings = bool(session.get("otp_setup_from_settings"))
        finish_setup_requested = request.method == "POST" and request.form.get("finish_setup") == "1"

        if finish_setup_requested and current_user.is_authenticated and session.get("pending_setup_recovery_codes"):
            target = "account_settings" if from_settings else "index"
            session.pop("pending_setup_recovery_codes", None)
            session.pop("pending_otp_user_id", None)
            session.pop("pending_otp_remember", None)
            session.pop("otp_setup_from_settings", None)
            return redirect(url_for(target))

        if not pending_user_id:
            flash("Сначала войдите в систему.", "danger")
            return redirect(url_for("login"))

        user_row = get_user_by_id(pending_user_id)
        if not user_row:
            session.pop("pending_otp_user_id", None)
            session.pop("pending_otp_remember", None)
            session.pop("otp_setup_from_settings", None)
            flash("Пользователь не найден.", "danger")
            return redirect(url_for("login"))

        otp_secret = user_row.get("otp_secret") or generate_totp_secret()
        if not user_row.get("otp_secret"):
            set_user_otp(user_row["id"], otp_secret, otp_enabled=False)
            user_row = get_user_by_id(pending_user_id)

        app_name = "Shans App"
        otp_uri = (
            f"otpauth://totp/{app_name}:{user_row['username']}"
            f"?secret={otp_secret}&issuer={app_name}"
        )
        qr_url = f"https://quickchart.io/qr?text={otp_uri}&size=220"

        generated_codes = session.get("pending_setup_recovery_codes", [])

        if request.method == "POST":
            if finish_setup_requested:
                target = "account_settings" if from_settings else "index"
                session.pop("pending_setup_recovery_codes", None)
                session.pop("pending_otp_user_id", None)
                session.pop("pending_otp_remember", None)
                session.pop("otp_setup_from_settings", None)
                return redirect(url_for(target))

            otp_code = request.form.get("otp_code", "").strip()
            if not verify_totp_code(otp_secret, otp_code):
                flash("Код не подошёл. Проверьте код в Google Authenticator.", "danger")
                return render_template(
                    "setup_2fa.html",
                    user=user_row,
                    otp_secret=otp_secret,
                    otp_uri=otp_uri,
                    qr_url=qr_url,
                )

            set_user_otp(user_row["id"], otp_secret, otp_enabled=True)
            remember_me = bool(session.get("pending_otp_remember"))
            authenticated_user = load_user_from_db(user_row["id"])
            login_user(authenticated_user, remember=remember_me)
            log_audit(current_app, "user_login_2fa_setup_completed", username=user_row["username"])
            recovery_codes = [f"{secrets.token_hex(2)}-{secrets.token_hex(2)}".upper() for _ in range(8)]
            session["account_recovery_codes"] = recovery_codes
            session["pending_setup_recovery_codes"] = recovery_codes
            flash("Google Authenticator подключён. Сохраните резервные коды восстановления.", "success")
            return render_template(
                "setup_2fa.html",
                user=user_row,
                otp_secret=otp_secret,
                otp_uri=otp_uri,
                qr_url=qr_url,
                setup_complete=True,
                recovery_codes=recovery_codes,
                from_settings=from_settings,
            )

        return render_template(
            "setup_2fa.html",
            user=user_row,
            otp_secret=otp_secret,
            otp_uri=otp_uri,
            qr_url=qr_url,
            setup_complete=False,
            recovery_codes=generated_codes,
            from_settings=from_settings,
        )

    @app.route("/account/settings/2fa/setup", methods=["POST"])
    @login_required
    def start_account_2fa_setup():
        user_row = get_user_by_id(current_user.id)
        if not user_row:
            flash("Пользователь не найден.", "danger")
            return redirect(url_for("account_settings"))

        pending_secret = generate_totp_secret()
        set_user_otp(current_user.id, pending_secret, otp_enabled=False)
        session["pending_otp_user_id"] = current_user.id
        session["pending_otp_remember"] = 1 if current_user.is_authenticated else 0
        session["otp_setup_from_settings"] = 1
        return redirect(url_for("setup_2fa"))

    @app.route("/account/settings")
    @login_required
    def account_settings():
        user_row = get_user_by_id(current_user.id)
        login_history = session.get("account_login_history", [])
        normalized_history = []
        for item in login_history:
            if not isinstance(item, dict):
                continue
            history_item = dict(item)
            if not history_item.get("session_key"):
                history_item["session_key"] = secrets.token_hex(12)
            normalized_history.append(history_item)
        login_history = normalized_history
        device_name, browser_name = _detect_device_and_browser(request.headers.get("User-Agent"))
        current_entry = {
            "session_key": session.get("account_current_session_key") or secrets.token_hex(12),
            "device": device_name,
            "browser": browser_name,
            "ip": request.headers.get("X-Forwarded-For", request.remote_addr or "—"),
            "date": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "current": True,
        }
        session["account_current_session_key"] = current_entry["session_key"]
        login_history = [current_entry] + [
            {**item, "current": False}
            for item in login_history[:9]
            if item.get("session_key") != current_entry["session_key"]
        ]
        session["account_login_history"] = login_history
        recovery_codes = session.get("account_recovery_codes", [])
        return render_template(
            "account_settings.html",
            avatar_letter=_build_avatar_letter(current_user.username),
            avatar_url=_build_avatar_url(user_row),
            otp_enabled=bool(user_row and user_row.get("otp_enabled")),
            login_history=login_history,
            recovery_codes=recovery_codes[:8],
        )

    @app.route("/account/settings/avatar", methods=["POST"])
    @login_required
    def update_account_avatar():
        cropped_avatar_data = request.form.get("cropped_avatar_data", "").strip()
        avatar_file = request.files.get("avatar_file")
        file_ext = ""

        if cropped_avatar_data.startswith("data:image/"):
            try:
                header, encoded = cropped_avatar_data.split(",", 1)
                if "image/png" not in header and "image/jpeg" not in header and "image/webp" not in header:
                    raise ValueError("unsupported format")
                raw_bytes = base64.b64decode(encoded)
                if len(raw_bytes) > 3 * 1024 * 1024:
                    flash("Файл слишком большой. Максимум 3 МБ.", "danger")
                    return redirect(url_for("account_settings"))
                file_ext = ".png" if "image/png" in header else ".jpg"
            except Exception:
                flash("Не удалось обработать обрезанное изображение.", "danger")
                return redirect(url_for("account_settings"))
        else:
            if not avatar_file or not avatar_file.filename:
                flash("Выберите изображение для аватарки.", "danger")
                return redirect(url_for("account_settings"))

            original_filename = secure_filename(avatar_file.filename)
            _, file_ext = os.path.splitext(original_filename.lower())
            allowed_ext = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
            if file_ext not in allowed_ext:
                flash("Разрешены только PNG/JPG/JPEG/WEBP/GIF.", "danger")
                return redirect(url_for("account_settings"))

            avatar_file.seek(0, os.SEEK_END)
            file_size = avatar_file.tell()
            avatar_file.seek(0)
            if file_size > 3 * 1024 * 1024:
                flash("Файл слишком большой. Максимум 3 МБ.", "danger")
                return redirect(url_for("account_settings"))

        upload_dir = os.path.join(current_app.static_folder, "uploads", "avatars")
        os.makedirs(upload_dir, exist_ok=True)
        new_filename = f"user_{current_user.id}_{secrets.token_hex(8)}{file_ext}"
        target_path = os.path.join(upload_dir, new_filename)
        if cropped_avatar_data.startswith("data:image/"):
            with open(target_path, "wb") as image_file:
                image_file.write(raw_bytes)
        else:
            avatar_file.save(target_path)

        user_row = get_user_by_id(current_user.id)
        previous_filename = (user_row.get("avatar_filename") if user_row else "") or ""
        if previous_filename and previous_filename != new_filename:
            previous_path = os.path.join(upload_dir, previous_filename)
            if os.path.exists(previous_path):
                os.remove(previous_path)

        set_user_avatar_filename(current_user.id, new_filename)
        flash("Аватарка обновлена.", "success")
        return redirect(url_for("account_settings"))

    @app.route("/account/settings/avatar/remove", methods=["POST"])
    @login_required
    def remove_account_avatar():
        upload_dir = os.path.join(current_app.static_folder, "uploads", "avatars")
        user_row = get_user_by_id(current_user.id)
        previous_filename = (user_row.get("avatar_filename") if user_row else "") or ""
        if previous_filename:
            previous_path = os.path.join(upload_dir, previous_filename)
            if os.path.exists(previous_path):
                os.remove(previous_path)
        set_user_avatar_filename(current_user.id, "")
        flash("Аватарка удалена.", "success")
        return redirect(url_for("account_settings"))

    @app.route("/account/settings/password", methods=["POST"])
    @login_required
    def update_account_password():
        current_password = request.form.get("current_password", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        user_row = get_user_by_id(current_user.id)
        if not user_row:
            flash("Пользователь не найден.", "danger")
            return redirect(url_for("account_settings"))

        if not check_password_hash(user_row.get("password_hash", ""), current_password):
            flash("Текущий пароль указан неверно.", "danger")
            return redirect(url_for("account_settings"))

        if len(new_password) < 8:
            flash("Новый пароль должен содержать минимум 8 символов.", "danger")
            return redirect(url_for("account_settings"))

        if new_password != confirm_password:
            flash("Подтверждение пароля не совпадает.", "danger")
            return redirect(url_for("account_settings"))

        if current_password == new_password:
            flash("Новый пароль должен отличаться от текущего.", "danger")
            return redirect(url_for("account_settings"))

        set_user_password_hash(current_user.id, generate_password_hash(new_password))
        flash("Пароль успешно обновлён.", "success")
        return redirect(url_for("account_settings"))

    @app.route("/account/settings/2fa/disable", methods=["POST"])
    @login_required
    def disable_account_2fa():
        user_row = get_user_by_id(current_user.id)
        otp_code = request.form.get("otp_code", "").strip()
        if not user_row:
            flash("Пользователь не найден.", "danger")
            return redirect(url_for("account_settings"))

        if user_row.get("otp_enabled") and not verify_totp_code(user_row.get("otp_secret"), otp_code):
            flash("Неверный код Google Authenticator.", "danger")
            return redirect(url_for("account_settings"))

        set_user_otp(current_user.id, "", otp_enabled=False)
        flash("2FA отключена для аккаунта.", "success")
        return redirect(url_for("account_settings"))

    @app.route("/account/settings/logout-all", methods=["POST"])
    @login_required
    def logout_all_devices():
        username = current_user.username
        session.clear()
        logout_user()
        log_audit(current_app, "user_logout_all_devices", username=username)
        flash("Вы вышли из аккаунта на текущем устройстве. Рекомендуем сменить пароль для завершения остальных сессий.", "success")
        return redirect(url_for("login"))

    @app.route("/account/settings/logout-device", methods=["POST"])
    @login_required
    def logout_device_session():
        target_session_key = request.form.get("session_key", "").strip()
        if not target_session_key:
            flash("Сессия не найдена.", "danger")
            return redirect(url_for("account_settings"))

        current_session_key = session.get("account_current_session_key")
        if target_session_key == current_session_key:
            username = current_user.username
            session.clear()
            logout_user()
            log_audit(current_app, "user_logout_device_session_current", username=username)
            flash("Текущая сессия завершена.", "success")
            return redirect(url_for("login"))

        login_history = session.get("account_login_history", [])
        filtered_history = [
            item for item in login_history
            if isinstance(item, dict) and item.get("session_key") != target_session_key
        ]
        if len(filtered_history) == len(login_history):
            flash("Выбранная сессия не найдена.", "warning")
        else:
            session["account_login_history"] = filtered_history
            flash("Выбранная сессия завершена.", "success")
        return redirect(url_for("account_settings"))

    @app.route("/logout")
    @login_required
    def logout():
        username = current_user.username
        logout_user()
        log_audit(current_app, "user_logout", username=username)
        return redirect(url_for("login"))

    # =========================================================
    # BUDGET
    # =========================================================

    @app.route("/budget")
    @login_required
    def budget():
        current_month = _get_current_month_name()
        summary = get_budget_summary(current_month)

        current_balance = get_balance_for_month(current_month)
        if current_balance is None:
            current_balance = get_current_balance()

        return render_template(
            "budget.html",
            current_month=current_month,
            summary=summary,
            current_balance=current_balance,
        )

    @app.route("/budget/balance")
    @login_required
    def budget_balance():
        history = get_balance_history()

        return render_template(
            "budget_balance.html",
            history=history,
        )

    @app.route("/budget/manage", methods=["GET", "POST"])
    @login_required
    def budget_manage():
        if request.method == "POST":
            form_type = request.form.get("form_type", "").strip()

            if form_type == "add_entry":
                month_name = request.form.get("month_name", "").strip()
                entry_type = request.form.get("entry_type", "").strip()
                amount_raw = request.form.get("amount", "").strip()
                category = _resolve_budget_category(request.form)

                if not month_name or not entry_type or not amount_raw:
                    log_invalid_form(current_app, "budget_add_entry", "missing_required_fields")
                    flash("Заполни все обязательные поля.", "warning")
                    return redirect(url_for("budget_manage"))

                if not category:
                    flash("Если выбрано 'Другое', укажи свою категорию.", "warning")
                    return redirect(url_for("budget_manage"))

                try:
                    amount_value = int(amount_raw)
                except ValueError:
                    flash("Сумма должна быть целым числом.", "danger")
                    return redirect(url_for("budget_manage"))

                if amount_value < 0:
                    flash("Сумма не может быть отрицательной.", "warning")
                    return redirect(url_for("budget_manage"))

                create_budget_entry(
                    entry_type=entry_type,
                    month_name=month_name,
                    category=category,
                    amount=amount_value,
                )
                log_audit(
                    current_app,
                    "budget_entry_created",
                    month=month_name,
                    entry_type=entry_type,
                    category=category,
                    amount=amount_value,
                )
                flash("Запись успешно добавлена.", "success")
                return redirect(url_for("budget_manage"))

            if form_type == "set_balance":
                balance_raw = request.form.get("current_balance", "").strip()

                if not balance_raw:
                    flash("Введите текущий баланс.", "warning")
                    return redirect(url_for("budget_manage"))

                try:
                    balance_value = int(balance_raw)
                except ValueError:
                    flash("Баланс должен быть целым числом.", "danger")
                    return redirect(url_for("budget_manage"))

                if balance_value < 0:
                    flash("Баланс не может быть отрицательным.", "warning")
                    return redirect(url_for("budget_manage"))

                current_month = _get_current_month_name()

                set_current_balance(balance_value)
                save_balance_history(current_month, balance_value)

                flash("Текущий баланс обновлён.", "success")
                return redirect(url_for("budget_manage"))

        month_filter = request.args.get("month", "").strip()
        type_filter = request.args.get("type_filter", "").strip()
        category_filter = request.args.get("category_filter", "").strip()
        sort_by = request.args.get("sort_by", "newest").strip()

        entries = get_all_budget_entries(
            month_filter=month_filter,
            type_filter=type_filter,
            category_filter=category_filter,
            sort_by=sort_by,
        )

        all_entries = get_all_budget_entries(sort_by="newest")
        categories = _collect_budget_categories(all_entries)

        current_month = _get_current_month_name()
        current_balance = get_balance_for_month(current_month)
        if current_balance is None:
            current_balance = get_current_balance()

        return render_template(
            "budget_manage.html",
            months=MONTHS,
            categories=categories,
            base_categories=BASE_CATEGORIES,
            current_balance=current_balance,
            entries=entries,
            month_filter=month_filter,
            type_filter=type_filter,
            category_filter=category_filter,
            sort_by=sort_by,
        )

    @app.route("/budget/report")
    @login_required
    def budget_report():
        current_month = _get_current_month_name()

        selected_month = request.args.get("month", current_month).strip()
        type_filter = request.args.get("type_filter", "").strip()
        category_filter = request.args.get("category_filter", "").strip()

        entries = get_all_budget_entries(
            month_filter=selected_month,
            type_filter=type_filter,
            category_filter=category_filter,
            sort_by="newest",
        )
        all_entries = get_all_budget_entries(sort_by="newest")
        categories = _collect_budget_categories(all_entries)

        summary = get_budget_summary(selected_month)

        return render_template(
            "budget_report.html",
            entries=entries,
            summary=summary,
            selected_month=selected_month,
            type_filter=type_filter,
            category_filter=category_filter,
            months=MONTHS,
            categories=categories,
        )

    @app.route("/budget/edit/<int:entry_id>", methods=["GET", "POST"])
    @login_required
    def budget_edit(entry_id):
        entry = get_budget_entry_by_id(entry_id)
        if not entry:
            flash("Запись не найдена.", "danger")
            return redirect(url_for("budget_manage"))

        all_entries = get_all_budget_entries(sort_by="newest")
        categories = _collect_budget_categories(all_entries)

        if request.method == "POST":
            entry_type = request.form.get("entry_type", "").strip()
            month_name = request.form.get("month_name", "").strip()
            amount_raw = request.form.get("amount", "").strip()
            category = _resolve_budget_category(request.form)

            if not entry_type or not month_name or not amount_raw:
                flash("Заполни все обязательные поля.", "warning")
                return redirect(url_for("budget_edit", entry_id=entry_id))

            if not category:
                flash("Если выбрано 'Другое', укажи свою категорию.", "warning")
                return redirect(url_for("budget_edit", entry_id=entry_id))

            try:
                amount_value = int(amount_raw)
            except ValueError:
                flash("Сумма должна быть целым числом.", "danger")
                return redirect(url_for("budget_edit", entry_id=entry_id))

            if amount_value < 0:
                flash("Сумма не может быть отрицательной.", "warning")
                return redirect(url_for("budget_edit", entry_id=entry_id))

            update_budget_entry(
                entry_id=entry_id,
                entry_type=entry_type,
                month_name=month_name,
                category=category,
                amount=amount_value,
            )
            log_audit(current_app, "budget_entry_updated", entry_id=entry_id)
            flash("Запись успешно обновлена.", "success")
            return redirect(url_for("budget_manage"))

        return render_template(
            "budget_edit.html",
            entry=entry,
            months=MONTHS,
            categories=categories,
            base_categories=BASE_CATEGORIES,
        )

    @app.route("/budget/delete/<int:entry_id>", methods=["POST"])
    @login_required
    def budget_delete(entry_id):
        delete_budget_entry(entry_id)
        log_audit(current_app, "budget_entry_deleted", entry_id=entry_id)
        flash("Запись успешно удалена.", "success")
        return redirect(url_for("budget_manage"))

    @app.route("/budget/delete-selected", methods=["POST"])
    @login_required
    def budget_delete_selected():
        selected_ids = _parse_selected_ids(request.form.get("selected_ids", ""))
        if not selected_ids:
            flash("Выбери хотя бы одну запись для удаления.", "warning")
            return redirect(url_for("budget_manage"))

        for entry_id in selected_ids:
            delete_budget_entry(entry_id)
        flash(f"Удалено записей: {len(selected_ids)}.", "success")
        return redirect(url_for("budget_manage"))

    @app.route("/budget/delete-all", methods=["POST"])
    @login_required
    def budget_delete_all():
        entries = get_all_budget_entries()
        if not entries:
            flash("В таблице бюджета нет записей для удаления.", "warning")
            return redirect(url_for("budget_manage"))

        for entry in entries:
            delete_budget_entry(entry["id"])
        flash("Все записи бюджета удалены.", "success")
        return redirect(url_for("budget_manage"))

    @app.route("/budget/export")
    @login_required
    def budget_export():
        current_month = _get_current_month_name()

        selected_month = request.args.get("month", current_month).strip()
        type_filter = request.args.get("type_filter", "").strip()
        category_filter = request.args.get("category_filter", "").strip()

        entries = get_all_budget_entries(
            month_filter=selected_month,
            type_filter=type_filter,
            category_filter=category_filter,
            sort_by="newest",
        )
        summary = get_budget_summary(selected_month)

        current_balance = get_balance_for_month(selected_month)
        if current_balance is None:
            current_balance = get_current_balance()

        excel_file = build_budget_excel(
            entries=entries,
            summary=summary,
            current_balance=current_balance,
            selected_month=selected_month,
        )

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f"budget_{selected_month}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # =========================================================
    # OTHER SECTIONS
    # =========================================================

    @app.route("/shootings")
    @login_required
    def shootings():
        return redirect(url_for("shootings_upcoming"))


    @app.route("/shootings/add", methods=["GET", "POST"])
    @login_required
    def shootings_add():
        form_data = {
            "project_name": "",
            "client_name": "",
            "shooting_date": "",
            "shooting_time": "",
            "duration_hours": "1",
            "phone": "",
            "price": "",
            "prepayment": "",
            "notes": "",
        }

        if request.method == "POST":
            project_name = request.form.get("project_name", "").strip()
            client_name = request.form.get("client_name", "").strip()
            shooting_date = request.form.get("shooting_date", "").strip()
            shooting_time = request.form.get("shooting_time", "").strip()
            duration_hours = request.form.get("duration_hours", "1").strip()
            phone = request.form.get("phone", "").strip()
            price = request.form.get("price", "0").strip()
            prepayment = request.form.get("prepayment", "0").strip()
            notes = request.form.get("notes", "").strip()

            form_data = {
                "project_name": project_name,
                "client_name": client_name,
                "shooting_date": shooting_date,
                "shooting_time": shooting_time,
                "duration_hours": duration_hours,
                "phone": phone,
                "price": price,
                "prepayment": prepayment,
                "notes": notes,
            }

            if not project_name or not client_name or not shooting_date:
                flash("Заполните обязательные поля: проект, клиент и дата.", "danger")
                return render_template("shootings_add.html", form_data=form_data, active_tab="add")

            try:
                duration_hours = float(duration_hours) if duration_hours else 1
                price = float(price) if price else 0
                prepayment = float(prepayment) if prepayment else 0
            except ValueError:
                flash("Часы, стоимость и предоплата должны быть числами.", "danger")
                return render_template("shootings_add.html", form_data=form_data, active_tab="add")
            if not _is_valid_contact_value(phone):
                flash("Укажи корректный контакт: номер телефона или ссылку.", "danger")
                return render_template("shootings_add.html", form_data=form_data, active_tab="add")

            shooting_id = create_shooting(
                project_name=project_name,
                client_name=client_name,
                shooting_date=shooting_date,
                shooting_time=shooting_time,
                duration_hours=duration_hours,
                phone=phone,
                price=price,
                prepayment=prepayment,
                notes=notes,
            )

            if shooting_id:
                upsert_task_for_shooting(
                    shooting_id=shooting_id,
                    project_name=project_name,
                    client_name=client_name,
                    shooting_date=shooting_date,
                    shooting_time=shooting_time,
                    duration_hours=duration_hours,
                    phone=phone,
                    price=price,
                    prepayment=prepayment,
                    notes=notes,
                )
                log_audit(
                    current_app,
                    "shooting_created",
                    shooting_id=shooting_id,
                    project_name=project_name,
                )

            flash("Съёмка успешно добавлена и появилась в графике.", "success")
            return redirect(url_for("shootings_upcoming"))

        return render_template("shootings_add.html", form_data=form_data, active_tab="add")


    @app.route("/shootings/upcoming")
    @login_required
    def shootings_upcoming():
        shootings = get_upcoming_shootings()
        shootings_count = get_shootings_count()
        nearest_shooting = get_nearest_shooting()

        return render_template(
            "shootings_upcoming.html",
            shootings=shootings,
            shootings_count=shootings_count,
            nearest_shooting=nearest_shooting,
            active_tab="upcoming",
        )


    @app.route("/shootings/export/<string:scope>")
    @login_required
    def shootings_export(scope):
        if scope == "archive":
            shootings = get_archived_shootings()
            report_title = "Архив съёмок"
            file_name = "shootings_archive.xlsx"
        else:
            shootings = get_upcoming_shootings()
            report_title = "Забронированные съёмки"
            file_name = "shootings_upcoming.xlsx"

        excel_file = build_shootings_excel(
            shootings=shootings,
            report_title=report_title,
        )

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    @app.route("/shootings/archive")
    @login_required
    def shootings_archive():
        shootings = get_archived_shootings()

        return render_template(
            "shootings_archive.html",
            shootings=shootings,
            active_tab="archive",
        )


    @app.route("/shootings/<int:shooting_id>")
    @login_required
    def shooting_detail(shooting_id):
        shooting = get_shooting_by_id(shooting_id)

        if not shooting:
            flash("Съёмка не найдена.", "danger")
            return redirect(url_for("shootings_upcoming"))

        return render_template("shooting_detail.html", shooting=shooting)


    @app.route("/shootings/<int:shooting_id>/edit", methods=["GET", "POST"])
    @login_required
    def shooting_edit(shooting_id):
        shooting = get_shooting_by_id(shooting_id)

        if not shooting:
            flash("Съёмка не найдена.", "danger")
            return redirect(url_for("shootings_upcoming"))

        if request.method == "POST":
            project_name = request.form.get("project_name", "").strip()
            client_name = request.form.get("client_name", "").strip()
            shooting_date = request.form.get("shooting_date", "").strip()
            shooting_time = request.form.get("shooting_time", "").strip()
            duration_hours = request.form.get("duration_hours", "1").strip()
            phone = request.form.get("phone", "").strip()
            price = request.form.get("price", "0").strip()
            prepayment = request.form.get("prepayment", "0").strip()
            notes = request.form.get("notes", "").strip()

            if not project_name or not client_name or not shooting_date:
                flash("Заполните обязательные поля: проект, клиент и дата.", "danger")
                return redirect(url_for("shooting_edit", shooting_id=shooting_id))

            try:
                duration_hours = float(duration_hours) if duration_hours else 1
                price = float(price) if price else 0
                prepayment = float(prepayment) if prepayment else 0
            except ValueError:
                flash("Часы, стоимость и предоплата должны быть числами.", "danger")
                return redirect(url_for("shooting_edit", shooting_id=shooting_id))
            if not _is_valid_contact_value(phone):
                flash("Укажи корректный контакт: номер телефона или ссылку.", "danger")
                return redirect(url_for("shooting_edit", shooting_id=shooting_id))

            update_shooting(
                shooting_id=shooting_id,
                project_name=project_name,
                client_name=client_name,
                shooting_date=shooting_date,
                shooting_time=shooting_time,
                duration_hours=duration_hours,
                phone=phone,
                price=price,
                prepayment=prepayment,
                notes=notes,
            )

            upsert_task_for_shooting(
                shooting_id=shooting_id,
                project_name=project_name,
                client_name=client_name,
                shooting_date=shooting_date,
                shooting_time=shooting_time,
                duration_hours=duration_hours,
                phone=phone,
                price=price,
                prepayment=prepayment,
                notes=notes,
            )
            log_audit(current_app, "shooting_updated", shooting_id=shooting_id)

            flash("Съёмка обновлена и синхронизирована с графиком.", "success")
            return redirect(url_for("shooting_detail", shooting_id=shooting_id))

        return render_template("shooting_edit.html", shooting=shooting)


    @app.route("/shootings/<int:shooting_id>/delete", methods=["POST"])
    @login_required
    def shooting_delete(shooting_id):
        shooting = get_shooting_by_id(shooting_id)

        if not shooting:
            flash("Съёмка не найдена.", "danger")
            return redirect(url_for("shootings_upcoming"))

        delete_task_for_shooting(shooting_id)
        delete_shooting(shooting_id)
        log_audit(current_app, "shooting_deleted", shooting_id=shooting_id)
        flash("Съёмка удалена из раздела съёмок и из графика.", "success")
        return redirect(url_for("shootings_upcoming"))

    @app.route("/shootings/upcoming/delete-selected", methods=["POST"])
    @login_required
    def shootings_upcoming_delete_selected():
        selected_ids = _parse_selected_ids(request.form.get("selected_ids", ""))
        if not selected_ids:
            flash("Выбери хотя бы одну съёмку.", "warning")
            return redirect(url_for("shootings_upcoming"))

        deleted_count = 0
        for shooting_id in selected_ids:
            shooting = get_shooting_by_id(shooting_id)
            shooting_data = dict(shooting) if shooting else None
            if not shooting_data or shooting_data.get("is_archive"):
                continue
            delete_task_for_shooting(shooting_id)
            delete_shooting(shooting_id)
            deleted_count += 1

        if deleted_count == 0:
            flash("Нет подходящих съёмок для удаления.", "warning")
        else:
            flash(f"Удалено съёмок: {deleted_count}.", "success")
        return redirect(url_for("shootings_upcoming"))

    @app.route("/shootings/upcoming/delete-all", methods=["POST"])
    @login_required
    def shootings_upcoming_delete_all():
        shootings = get_upcoming_shootings()
        if not shootings:
            flash("Нет забронированных съёмок для удаления.", "warning")
            return redirect(url_for("shootings_upcoming"))

        for shooting in shootings:
            shooting_id = int(shooting["id"])
            delete_task_for_shooting(shooting_id)
            delete_shooting(shooting_id)

        flash("Все забронированные съёмки удалены.", "success")
        return redirect(url_for("shootings_upcoming"))

    @app.route("/shootings/archive/delete-selected", methods=["POST"])
    @login_required
    def shootings_archive_delete_selected():
        selected_ids = _parse_selected_ids(request.form.get("selected_ids", ""))
        if not selected_ids:
            flash("Выбери хотя бы одну архивную съёмку.", "warning")
            return redirect(url_for("shootings_archive"))

        deleted_count = 0
        for shooting_id in selected_ids:
            shooting = get_shooting_by_id(shooting_id)
            shooting_data = dict(shooting) if shooting else None
            if not shooting_data or not shooting_data.get("is_archive"):
                continue
            delete_task_for_shooting(shooting_id)
            delete_shooting(shooting_id)
            deleted_count += 1

        if deleted_count == 0:
            flash("Нет подходящих архивных съёмок для удаления.", "warning")
        else:
            flash(f"Удалено архивных съёмок: {deleted_count}.", "success")
        return redirect(url_for("shootings_archive"))

    @app.route("/shootings/archive/delete-all", methods=["POST"])
    @login_required
    def shootings_archive_delete_all():
        shootings = get_archived_shootings()
        if not shootings:
            flash("Нет архивных съёмок для удаления.", "warning")
            return redirect(url_for("shootings_archive"))

        for shooting in shootings:
            shooting_id = int(shooting["id"])
            delete_task_for_shooting(shooting_id)
            delete_shooting(shooting_id)

        flash("Все архивные съёмки удалены.", "success")
        return redirect(url_for("shootings_archive"))

    @app.route("/scenarios")
    @login_required
    def scenarios():
        return redirect(url_for("scenarios_upcoming"))

    @app.route("/scenarios/add", methods=["GET", "POST"])
    @login_required
    def scenarios_add():
        form_data = {"title": "", "shooting_date": "", "scenario_text": "", "scenario_status": "in_progress"}

        if request.method == "POST":
            title = request.form.get("title", "").strip()
            shooting_date = request.form.get("shooting_date", "").strip()
            scenario_text = request.form.get("scenario_text", "").strip()
            scenario_status = request.form.get("scenario_status", "in_progress").strip()
            form_data = {
                "title": title,
                "shooting_date": shooting_date,
                "scenario_text": scenario_text,
                "scenario_status": scenario_status,
            }
            if scenario_status not in ["in_progress", "done"]:
                scenario_status = "in_progress"

            if not title or not shooting_date:
                flash("Заполните обязательные поля: название и дата съёмки.", "danger")
                return render_template("scenarios_add.html", form_data=form_data, active_tab="add")

            try:
                datetime.strptime(shooting_date, "%Y-%m-%d")
            except ValueError:
                flash("Укажи корректную дату съёмки.", "danger")
                return render_template("scenarios_add.html", form_data=form_data, active_tab="add")

            if len(scenario_text) > 700:
                flash("Текст сценария не должен превышать 700 символов.", "danger")
                return render_template("scenarios_add.html", form_data=form_data, active_tab="add")

            scenario_id = create_scenario(
                title=title,
                shooting_date=shooting_date,
                scenario_text=scenario_text,
                scenario_status=scenario_status,
            )
            upsert_task_for_scenario(
                scenario_id=scenario_id,
                title=title,
                shooting_date=shooting_date,
                scenario_text=scenario_text,
                scenario_status=scenario_status,
            )
            log_audit(current_app, "scenario_created", scenario_id=scenario_id, title=title)
            flash("Сценарий успешно добавлен и синхронизирован с графиком.", "success")
            return redirect(url_for("scenarios_upcoming"))

        return render_template("scenarios_add.html", form_data=form_data, active_tab="add")

    @app.route("/scenarios/upcoming")
    @login_required
    def scenarios_upcoming():
        scenarios = get_upcoming_scenarios()
        scenarios_count = get_scenarios_count()
        nearest_scenario = get_nearest_scenario()
        return render_template(
            "scenarios_upcoming.html",
            scenarios=scenarios,
            scenarios_count=scenarios_count,
            nearest_scenario=nearest_scenario,
            active_tab="upcoming",
        )

    @app.route("/scenarios/archive")
    @login_required
    def scenarios_archive():
        scenarios = get_archived_scenarios()
        return render_template("scenarios_archive.html", scenarios=scenarios, active_tab="archive")

    @app.route("/scenarios/export/<string:scope>")
    @login_required
    def scenarios_export(scope):
        if scope == "archive":
            scenarios = get_archived_scenarios()
            report_title = "Архив сценариев"
            file_name = "scenarios_archive.xlsx"
        else:
            scenarios = get_upcoming_scenarios()
            report_title = "Активные сценарии"
            file_name = "scenarios_active.xlsx"

        excel_file = build_scenarios_excel(scenarios=scenarios, report_title=report_title)
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/scenarios/<int:scenario_id>")
    @login_required
    def scenario_detail(scenario_id):
        scenario = get_scenario_by_id(scenario_id)
        if not scenario:
            flash("Сценарий не найден.", "danger")
            return redirect(url_for("scenarios_upcoming"))
        return render_template("scenario_detail.html", scenario=scenario)

    @app.route("/scenarios/<int:scenario_id>/edit", methods=["GET", "POST"])
    @login_required
    def scenario_edit(scenario_id):
        scenario = get_scenario_by_id(scenario_id)
        if not scenario:
            flash("Сценарий не найден.", "danger")
            return redirect(url_for("scenarios_upcoming"))

        if request.method == "POST":
            title = request.form.get("title", "").strip()
            shooting_date = request.form.get("shooting_date", "").strip()
            scenario_text = request.form.get("scenario_text", "").strip()
            scenario_status = request.form.get("scenario_status", "in_progress").strip()
            if scenario_status not in ["in_progress", "done"]:
                scenario_status = "in_progress"

            if not title or not shooting_date:
                flash("Заполните обязательные поля: название и дата съёмки.", "danger")
                return redirect(url_for("scenario_edit", scenario_id=scenario_id))

            try:
                datetime.strptime(shooting_date, "%Y-%m-%d")
            except ValueError:
                flash("Укажи корректную дату съёмки.", "danger")
                return redirect(url_for("scenario_edit", scenario_id=scenario_id))

            if len(scenario_text) > 700:
                flash("Текст сценария не должен превышать 700 символов.", "danger")
                return redirect(url_for("scenario_edit", scenario_id=scenario_id))

            update_scenario(
                scenario_id=scenario_id,
                title=title,
                shooting_date=shooting_date,
                scenario_text=scenario_text,
                scenario_status=scenario_status,
            )
            upsert_task_for_scenario(
                scenario_id=scenario_id,
                title=title,
                shooting_date=shooting_date,
                scenario_text=scenario_text,
                scenario_status=scenario_status,
            )
            log_audit(current_app, "scenario_updated", scenario_id=scenario_id)
            flash("Сценарий обновлён и синхронизирован с графиком.", "success")
            return redirect(url_for("scenario_detail", scenario_id=scenario_id))

        return render_template("scenario_edit.html", scenario=scenario)

    @app.route("/scenarios/<int:scenario_id>/toggle-status", methods=["POST"])
    @login_required
    def scenario_toggle_status(scenario_id):
        scenario = get_scenario_by_id(scenario_id)
        if not scenario:
            flash("Сценарий не найден.", "danger")
            return redirect(url_for("scenarios_upcoming"))

        toggle_scenario_status(scenario_id)
        updated_scenario = get_scenario_by_id(scenario_id)
        if updated_scenario:
            upsert_task_for_scenario(
                scenario_id=scenario_id,
                title=updated_scenario.get("title", ""),
                shooting_date=updated_scenario.get("shooting_date", ""),
                scenario_text=updated_scenario.get("scenario_text", ""),
                scenario_status=updated_scenario.get("scenario_status", "in_progress"),
            )
        log_audit(current_app, "scenario_status_toggled", scenario_id=scenario_id)
        flash("Статус сценария обновлён.", "success")
        if updated_scenario and updated_scenario.get("is_archive"):
            return redirect(url_for("scenarios_archive"))
        return redirect(url_for("scenarios_upcoming"))

    @app.route("/scenarios/<int:scenario_id>/delete", methods=["POST"])
    @login_required
    def scenario_delete(scenario_id):
        scenario = get_scenario_by_id(scenario_id)
        if not scenario:
            flash("Сценарий не найден.", "danger")
            return redirect(url_for("scenarios_upcoming"))

        delete_task_for_scenario(scenario_id)
        delete_scenario(scenario_id)
        log_audit(current_app, "scenario_deleted", scenario_id=scenario_id)
        flash("Сценарий удалён из раздела сценариев и из графика.", "success")
        return redirect(url_for("scenarios_upcoming"))

    @app.route("/scenarios/upcoming/delete-selected", methods=["POST"])
    @login_required
    def scenarios_upcoming_delete_selected():
        selected_ids = _parse_selected_ids(request.form.get("selected_ids", ""))
        if not selected_ids:
            flash("Выбери хотя бы один сценарий.", "warning")
            return redirect(url_for("scenarios_upcoming"))

        deleted_count = 0
        for scenario_id in selected_ids:
            scenario = get_scenario_by_id(scenario_id)
            scenario_data = dict(scenario) if scenario else None
            if not scenario_data or scenario_data.get("is_archive"):
                continue
            delete_task_for_scenario(scenario_id)
            delete_scenario(scenario_id)
            deleted_count += 1

        if deleted_count == 0:
            flash("Нет подходящих сценариев для удаления.", "warning")
        else:
            flash(f"Удалено сценариев: {deleted_count}.", "success")
        return redirect(url_for("scenarios_upcoming"))

    @app.route("/scenarios/upcoming/delete-all", methods=["POST"])
    @login_required
    def scenarios_upcoming_delete_all():
        scenarios = get_upcoming_scenarios()
        if not scenarios:
            flash("Нет активных сценариев для удаления.", "warning")
            return redirect(url_for("scenarios_upcoming"))

        for scenario in scenarios:
            scenario_id = int(scenario["id"])
            delete_task_for_scenario(scenario_id)
            delete_scenario(scenario_id)

        flash("Все активные сценарии удалены.", "success")
        return redirect(url_for("scenarios_upcoming"))

    @app.route("/scenarios/archive/delete-selected", methods=["POST"])
    @login_required
    def scenarios_archive_delete_selected():
        selected_ids = _parse_selected_ids(request.form.get("selected_ids", ""))
        if not selected_ids:
            flash("Выбери хотя бы один архивный сценарий.", "warning")
            return redirect(url_for("scenarios_archive"))

        deleted_count = 0
        for scenario_id in selected_ids:
            scenario = get_scenario_by_id(scenario_id)
            scenario_data = dict(scenario) if scenario else None
            if not scenario_data or not scenario_data.get("is_archive"):
                continue
            delete_task_for_scenario(scenario_id)
            delete_scenario(scenario_id)
            deleted_count += 1

        if deleted_count == 0:
            flash("Нет подходящих архивных сценариев для удаления.", "warning")
        else:
            flash(f"Удалено архивных сценариев: {deleted_count}.", "success")
        return redirect(url_for("scenarios_archive"))

    @app.route("/scenarios/archive/delete-all", methods=["POST"])
    @login_required
    def scenarios_archive_delete_all():
        scenarios = get_archived_scenarios()
        if not scenarios:
            flash("Нет архивных сценариев для удаления.", "warning")
            return redirect(url_for("scenarios_archive"))

        for scenario in scenarios:
            scenario_id = int(scenario["id"])
            delete_task_for_scenario(scenario_id)
            delete_scenario(scenario_id)

        flash("Все архивные сценарии удалены.", "success")
        return redirect(url_for("scenarios_archive"))

    # =========================================================
    # CAR
    # =========================================================

    @app.route("/car")
    @login_required
    def car():
        active_tab = request.args.get("tab", "planned")

        done_services = get_car_done_services()
        planned_services = get_car_planned_services()

        month_names = {
            "01": "январь",
            "02": "февраль",
            "03": "март",
            "04": "апрель",
            "05": "май",
            "06": "июнь",
            "07": "июль",
            "08": "август",
            "09": "сентябрь",
            "10": "октябрь",
            "11": "ноябрь",
            "12": "декабрь",
        }

        total_cost = 0
        max_mileage = 0

        prepared_done_services = []
        for item in done_services:
            row = dict(item)

            service_date = row.get("service_date")
            if service_date and "-" in str(service_date):
                parts = str(service_date).split("-")
                if len(parts) >= 2:
                    year = parts[0]
                    month = parts[1]
                    row["service_date_display"] = f"{month_names.get(month, month)} {year}"
                else:
                    row["service_date_display"] = service_date
            else:
                row["service_date_display"] = service_date or "—"

            cost_value = row.get("service_cost")
            mileage_value = row.get("mileage")

            try:
                if cost_value not in (None, ""):
                    total_cost += float(cost_value)
            except (ValueError, TypeError):
                pass

            try:
                if mileage_value not in (None, ""):
                    mileage_number = int(float(mileage_value))
                    if mileage_number > max_mileage:
                        max_mileage = mileage_number
            except (ValueError, TypeError):
                pass

            prepared_done_services.append(row)

        prepared_planned_services = [dict(item) for item in planned_services]

        return render_template(
            "car.html",
            active_tab=active_tab,
            done_services=prepared_done_services,
            planned_services=prepared_planned_services,
            total_cost=total_cost,
            max_mileage=max_mileage,
            planned_count=len(prepared_planned_services),
            done_count=len(prepared_done_services),
        )

    @app.route("/car/import-excel", methods=["POST"])
    @login_required
    def car_import_excel():
        excel_file = request.files.get("excel_file")
        if not excel_file or not excel_file.filename:
            flash("Выберите Excel-файл для загрузки.", "error")
            return redirect(url_for("car"))

        filename = excel_file.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
            flash("Поддерживаются только файлы .xlsx или .xlsm.", "error")
            return redirect(url_for("car"))

        try:
            done_services, planned_services = _parse_car_excel(excel_file)
            append_car_services(done_services=done_services, planned_services=planned_services)
            log_import_result(
                current_app,
                import_type="car",
                rows_added=len(done_services) + len(planned_services),
                errors_count=0,
                mode="append",
            )
        except ValueError as error:
            current_app.logger.warning("Car excel import validation error: %s", error)
            log_import_result(current_app, import_type="car", rows_added=0, errors_count=1, mode="append")
            flash(str(error), "error")
            return redirect(url_for("car"))
        except Exception:
            flash("Не удалось обработать Excel-файл. Проверьте формат данных.", "error")
            return redirect(url_for("car"))

        flash(
            (
                f"Импорт завершён: добавлено выполненных {len(done_services)}, "
                f"планируемых {len(planned_services)}."
            ),
            "success",
        )
        return redirect(url_for("car"))

    @app.route("/import-center", methods=["GET", "POST"])
    @login_required
    def import_center():
        if _is_mobile_request():
            flash("Раздел импорта доступен только на ПК-версии.", "error")
            return redirect(url_for("index"))

        access_granted = False

        if request.method == "POST":
            action = request.form.get("action", "").strip()

            if action == "unlock":
                password = request.form.get("password", "").strip()
                import_password = Config.IMPORT_CENTER_PASSWORD
                if import_password and hmac.compare_digest(password, import_password):
                    access_granted = True
                    flash("Доступ к разделу импорта открыт.", "success")
                else:
                    flash("Неверный пароль.", "error")
                return render_template("import_center.html", access_granted=access_granted)

            password = request.form.get("password", "").strip()
            import_password = Config.IMPORT_CENTER_PASSWORD
            if not import_password or not hmac.compare_digest(password, import_password):
                flash("Неверный пароль. Для каждого входа нужно подтверждение.", "error")
                return redirect(url_for("import_center"))

            target = request.form.get("target_section", "").strip()
            import_mode = request.form.get("import_mode", "append").strip()
            excel_file = request.files.get("excel_file")

            if not target:
                flash("Выберите раздел для импорта.", "error")
                return redirect(url_for("import_center"))

            if not excel_file or not excel_file.filename:
                flash("Выберите Excel-файл для загрузки.", "error")
                return redirect(url_for("import_center"))

            filename = excel_file.filename.lower()
            if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
                flash("Поддерживаются только файлы .xlsx или .xlsm.", "error")
                return redirect(url_for("import_center"))

            try:
                replace_mode = import_mode == "replace"
                if target in ("car_all", "car_done", "car_planned"):
                    done_services, planned_services = _parse_car_excel(excel_file)
                    if target == "car_done":
                        planned_services = []
                    elif target == "car_planned":
                        done_services = []
                    if replace_mode:
                        replace_car_services(done_services=done_services, planned_services=planned_services)
                    else:
                        append_car_services(done_services=done_services, planned_services=planned_services)
                    mode_text = "Таблица заменена" if replace_mode else "Добавлены записи"
                    log_import_result(
                        current_app,
                        import_type=target,
                        rows_added=len(done_services) + len(planned_services),
                        errors_count=0,
                        mode=import_mode,
                    )
                    flash(
                        (
                            f"{mode_text} в разделе Машина: выполненных {len(done_services)}, "
                            f"планируемых {len(planned_services)}."
                        ),
                        "success",
                    )
                elif target == "budget":
                    entries = _parse_budget_excel(excel_file)
                    if replace_mode:
                        replace_budget_entries(entries)
                        log_import_result(current_app, import_type="budget", rows_added=len(entries), errors_count=0, mode=import_mode)
                        flash(f"Таблица Бюджета заменена. Загружено записей: {len(entries)}.", "success")
                    else:
                        for entry in entries:
                            create_budget_entry(
                                entry_type=entry["entry_type"],
                                month_name=entry["month_name"],
                                category=entry["category"],
                                amount=entry["amount"],
                            )
                        log_import_result(current_app, import_type="budget", rows_added=len(entries), errors_count=0, mode=import_mode)
                        flash(f"Добавлено записей в Бюджет: {len(entries)}.", "success")
                elif target == "shootings":
                    shootings = _parse_shootings_excel(excel_file)
                    if replace_mode:
                        replace_shootings([])

                    for shooting in shootings:
                        shooting_id = create_shooting(
                            project_name=shooting["project_name"],
                            client_name=shooting["client_name"],
                            shooting_date=shooting["shooting_date"],
                            shooting_time=shooting["shooting_time"],
                            duration_hours=shooting["duration_hours"],
                            phone=shooting["phone"],
                            price=shooting["price"],
                            prepayment=shooting["prepayment"],
                            notes=shooting["notes"],
                        )
                        if shooting_id:
                            upsert_task_for_shooting(
                                shooting_id=shooting_id,
                                project_name=shooting["project_name"],
                                client_name=shooting["client_name"],
                                shooting_date=shooting["shooting_date"],
                                shooting_time=shooting["shooting_time"],
                                duration_hours=shooting["duration_hours"],
                                phone=shooting["phone"],
                                price=shooting["price"],
                                prepayment=shooting["prepayment"],
                                notes=shooting["notes"],
                            )
                    if replace_mode:
                        log_import_result(current_app, import_type="shootings", rows_added=len(shootings), errors_count=0, mode=import_mode)
                        flash(f"Таблица Съёмок заменена. Загружено записей: {len(shootings)}.", "success")
                    else:
                        log_import_result(current_app, import_type="shootings", rows_added=len(shootings), errors_count=0, mode=import_mode)
                        flash(f"Добавлено съёмок: {len(shootings)}.", "success")
                elif target == "schedule":
                    tasks = _parse_schedule_excel(excel_file)
                    if replace_mode:
                        replace_manual_schedule_tasks(tasks)
                        log_import_result(current_app, import_type="schedule", rows_added=len(tasks), errors_count=0, mode=import_mode)
                        flash(f"Таблица Графика заменена. Загружено задач: {len(tasks)}.", "success")
                    else:
                        for task in tasks:
                            create_task(
                                title=task["title"],
                                task_date=task["task_date"],
                                description=task["description"],
                                start_time=task["start_time"],
                                is_important=task["is_important"],
                                range_end_date=task["range_end_date"],
                                task_type=task["task_type"],
                                status=task["status"],
                            )
                        log_import_result(current_app, import_type="schedule", rows_added=len(tasks), errors_count=0, mode=import_mode)
                        flash(f"Добавлено задач в График: {len(tasks)}.", "success")
                else:
                    flash("Неизвестный раздел для импорта.", "error")
            except ValueError as error:
                current_app.logger.warning("Import-center validation error: target=%s mode=%s error=%s", target, import_mode, error)
                log_import_result(current_app, import_type=target or "unknown", rows_added=0, errors_count=1, mode=import_mode or "unknown")
                flash(str(error), "error")
            except Exception:
                current_app.logger.exception("Import-center unexpected error: target=%s mode=%s", target, import_mode)
                log_import_result(current_app, import_type=target or "unknown", rows_added=0, errors_count=1, mode=import_mode or "unknown")
                flash("Не удалось обработать файл. Проверьте формат данных.", "error")

            return redirect(url_for("import_center"))

        return render_template("import_center.html", access_granted=access_granted)

    @app.route("/log", methods=["GET", "POST"])
    @login_required
    def site_logs():
        if _is_mobile_request():
            flash("Раздел логов доступен только на ПК-версии.", "error")
            return redirect(url_for("index"))

        access_granted = False
        logs_payload = []
        available_logs = []
        logs_dir = os.path.abspath(os.path.join(current_app.root_path, "..", "logs"))

        if os.path.isdir(logs_dir):
            available_logs = sorted(
                [name for name in os.listdir(logs_dir) if name.endswith(".log")]
            )

        if request.method == "POST":
            action = request.form.get("action", "").strip()
            password = request.form.get("password", "").strip()
            import_password = Config.IMPORT_CENTER_PASSWORD

            if not import_password or not hmac.compare_digest(password, import_password):
                flash("Неверный пароль.", "error")
                return render_template(
                    "site_logs.html",
                    access_granted=False,
                    logs_payload=[],
                    available_logs=available_logs,
                    selected_log="",
                    log_lines_limit=500,
                )

            access_granted = True
            selected_log = request.form.get("log_name", "").strip()
            if action == "unlock":
                selected_log = available_logs[0] if available_logs else ""

            if selected_log and selected_log not in available_logs:
                flash("Запрошенный файл логов не найден.", "error")
                selected_log = available_logs[0] if available_logs else ""

            for log_name in available_logs:
                log_path = os.path.join(logs_dir, log_name)
                lines = []
                try:
                    with open(log_path, "r", encoding="utf-8", errors="replace") as source:
                        lines = source.readlines()[-500:]
                except OSError:
                    lines = ["Ошибка чтения файла логов."]
                logs_payload.append(
                    {
                        "name": log_name,
                        "content": "".join(lines).strip(),
                        "is_selected": log_name == selected_log,
                    }
                )

            if not logs_payload:
                flash("Файлы логов не найдены.", "error")

            return render_template(
                "site_logs.html",
                access_granted=access_granted,
                logs_payload=logs_payload,
                available_logs=available_logs,
                selected_log=selected_log,
                log_lines_limit=500,
            )

        return render_template(
            "site_logs.html",
            access_granted=access_granted,
            logs_payload=[],
            available_logs=available_logs,
            selected_log="",
            log_lines_limit=500,
        )

    @app.route("/car/manage", methods=["GET", "POST"])
    @login_required
    def car_manage():
        if request.method == "POST":
            service_name = request.form.get("service_name", "").strip()
            detail_description = request.form.get("detail_description", "").strip()
            work_kind = ""
            status = request.form.get("status", "").strip()
            period_type = request.form.get("period_type", "").strip()

            if not service_name:
                flash("Укажите наименование работы.", "error")
                return redirect(url_for("car_manage"))

            if status == "Выполнено":
                service_date = request.form.get("service_date", "").strip()
                mileage = request.form.get("mileage", "").strip()
                cost = request.form.get("cost", "").strip()
                mileage_value, mileage_error = _parse_non_negative_number(mileage, "Пробег")
                cost_value, cost_error = _parse_non_negative_number(cost, "Стоимость")

                if mileage_error or cost_error:
                    flash(mileage_error or cost_error, "error")
                    return redirect(url_for("car_manage"))

                create_car_done_service(
                    service_name=service_name,
                    service_cost=cost_value if cost_value is not None else "",
                    mileage=mileage_value if mileage_value is not None else "",
                    service_date=service_date,
                    detail_description=detail_description,
                    work_kind=work_kind,
                    period_type=period_type,
                )
                log_audit(current_app, "car_done_created", service_name=service_name)
                flash("Выполненная работа добавлена.", "success")
            else:
                create_car_planned_service(
                    service_name=service_name,
                    detail_description=detail_description,
                    work_kind=work_kind,
                    period_type=period_type,
                )
                log_audit(current_app, "car_planned_created", service_name=service_name)
                flash("Планируемая работа добавлена.", "success")

            return redirect(url_for("car"))

        return render_template("car_manage.html")

    @app.route("/car/done/edit/<int:service_id>", methods=["GET", "POST"])
    @login_required
    def car_done_edit(service_id):
        service = get_car_done_service_by_id(service_id)
        if not service:
            flash("Выполненная работа не найдена.", "error")
            return redirect(url_for("car", tab="done"))

        if request.method == "POST":
            service_name = request.form.get("service_name", "").strip()
            detail_description = request.form.get("detail_description", "").strip()
            service_date = request.form.get("service_date", "").strip()
            mileage = request.form.get("mileage", "").strip()
            cost = request.form.get("cost", "").strip()
            period_type = request.form.get("period_type", "").strip()
            mileage_value, mileage_error = _parse_non_negative_number(mileage, "Пробег")
            cost_value, cost_error = _parse_non_negative_number(cost, "Стоимость")

            if not service_name:
                flash("Укажите наименование работы.", "error")
                return redirect(url_for("car_done_edit", service_id=service_id))
            if mileage_error or cost_error:
                flash(mileage_error or cost_error, "error")
                return redirect(url_for("car_done_edit", service_id=service_id))

            update_car_done_service(
                service_id=service_id,
                service_name=service_name,
                service_cost=cost_value if cost_value is not None else "",
                mileage=mileage_value if mileage_value is not None else "",
                service_date=service_date,
                detail_description=detail_description,
                work_kind=service["work_kind"],
                period_type=period_type,
                status="Выполнено",
            )

            log_audit(current_app, "car_done_updated", service_id=service_id)
            flash("Выполненная работа обновлена.", "success")
            return redirect(url_for("car", tab="done"))

        return render_template("car_done_edit.html", service=service)

    @app.route("/car/planned/complete/<int:service_id>", methods=["GET", "POST"])
    @login_required
    def car_planned_complete(service_id):
        service = get_car_planned_service_by_id(service_id)
        if not service:
            flash("Планируемая работа не найдена.", "error")
            return redirect(url_for("car", tab="planned"))

        if request.method == "POST":
            service_name = request.form.get("service_name", "").strip()
            detail_description = request.form.get("detail_description", "").strip()
            period_type = request.form.get("period_type", "").strip()
            service_date = request.form.get("service_date", "").strip()
            mileage = request.form.get("mileage", "").strip()
            cost = request.form.get("cost", "").strip()
            mileage_value, mileage_error = _parse_non_negative_number(mileage, "Пробег")
            cost_value, cost_error = _parse_non_negative_number(cost, "Стоимость")

            if not service_name:
                flash("Укажите наименование работы.", "error")
                return redirect(url_for("car_planned_complete", service_id=service_id))
            if not service_date:
                flash("Укажите дату выполнения.", "error")
                return redirect(url_for("car_planned_complete", service_id=service_id))
            if mileage_error or cost_error:
                flash(mileage_error or cost_error, "error")
                return redirect(url_for("car_planned_complete", service_id=service_id))

            create_car_done_service(
                service_name=service_name,
                service_cost=cost_value if cost_value is not None else "",
                mileage=mileage_value if mileage_value is not None else "",
                service_date=service_date,
                detail_description=detail_description,
                work_kind=service["work_kind"],
                period_type=period_type,
            )

            delete_car_planned_service(service_id)
            log_audit(current_app, "car_planned_completed", service_id=service_id)
            flash("Работа перенесена в выполненные.", "success")
            return redirect(url_for("car", tab="done"))

        return render_template("car_planned_complete.html", service=service)

    @app.route("/car/planned/edit/<int:service_id>", methods=["GET", "POST"])
    @login_required
    def car_planned_edit(service_id):
        service = get_car_planned_service_by_id(service_id)
        if not service:
            flash("Планируемая работа не найдена.", "error")
            return redirect(url_for("car"))

        if request.method == "POST":
            service_name = request.form.get("service_name", "").strip()
            detail_description = request.form.get("detail_description", "").strip()
            period_type = request.form.get("period_type", "").strip()

            if not service_name:
                flash("Укажите наименование работы.", "error")
                return redirect(url_for("car_planned_edit", service_id=service_id))

            update_car_planned_service(
                service_id=service_id,
                service_name=service_name,
                detail_description=detail_description,
                work_kind=service["work_kind"],
                period_type=period_type,
            )

            log_audit(current_app, "car_planned_updated", service_id=service_id)
            flash("Планируемая работа обновлена.", "success")
            return redirect(url_for("car"))

        return render_template("car_planned_edit.html", service=service)

    @app.route("/car/done/delete/<int:service_id>", methods=["POST"])
    @login_required
    def car_done_delete(service_id):
        delete_car_done_service(service_id)
        log_audit(current_app, "car_done_deleted", service_id=service_id)
        flash("Выполненная работа удалена.", "success")
        return redirect(url_for("car", tab="done"))

    @app.route("/car/planned/delete/<int:service_id>", methods=["POST"])
    @login_required
    def car_planned_delete(service_id):
        delete_car_planned_service(service_id)
        log_audit(current_app, "car_planned_deleted", service_id=service_id)
        flash("Планируемая работа удалена.", "success")
        return redirect(url_for("car", tab="planned"))

    @app.route("/car/done/delete-selected", methods=["POST"])
    @login_required
    def car_done_delete_selected():
        selected_ids = _parse_selected_ids(request.form.get("selected_ids", ""))
        if not selected_ids:
            flash("Выбери хотя бы одну выполненную работу.", "warning")
            return redirect(url_for("car", tab="done"))

        for service_id in selected_ids:
            delete_car_done_service(service_id)
        flash(f"Удалено выполненных работ: {len(selected_ids)}.", "success")
        return redirect(url_for("car", tab="done"))

    @app.route("/car/done/delete-all", methods=["POST"])
    @login_required
    def car_done_delete_all():
        items = get_car_done_services()
        if not items:
            flash("Нет выполненных работ для удаления.", "warning")
            return redirect(url_for("car", tab="done"))

        for item in items:
            delete_car_done_service(item["id"])
        flash("Все выполненные работы удалены.", "success")
        return redirect(url_for("car", tab="done"))

    @app.route("/car/planned/delete-selected", methods=["POST"])
    @login_required
    def car_planned_delete_selected():
        selected_ids = _parse_selected_ids(request.form.get("selected_ids", ""))
        if not selected_ids:
            flash("Выбери хотя бы одну планируемую работу.", "warning")
            return redirect(url_for("car", tab="planned"))

        for service_id in selected_ids:
            delete_car_planned_service(service_id)
        flash(f"Удалено планируемых работ: {len(selected_ids)}.", "success")
        return redirect(url_for("car", tab="planned"))

    @app.route("/car/planned/delete-all", methods=["POST"])
    @login_required
    def car_planned_delete_all():
        items = get_car_planned_services()
        if not items:
            flash("Нет планируемых работ для удаления.", "warning")
            return redirect(url_for("car", tab="planned"))

        for item in items:
            delete_car_planned_service(item["id"])
        flash("Все планируемые работы удалены.", "success")
        return redirect(url_for("car", tab="planned"))

    @app.route("/car/notifications/archive/delete", methods=["POST"])
    @login_required
    def car_notification_archive_delete():
        notification_key = request.form.get("notification_key", "").strip()

        if not notification_key:
            flash("Не удалось удалить уведомление из архива.", "error")
            return redirect(url_for("car_notifications", filter="archive"))

        delete_archived_car_notification(notification_key)
        flash("Уведомление удалено из архива.", "success")
        return redirect(url_for("car_notifications", filter="archive"))

    @app.route("/car/notifications")
    @login_required
    def car_notifications():
        active_filter = request.args.get("filter", "need")
        notifications = _build_car_notifications()

        need_notifications = [item for item in notifications if item["status"] == "Нужна замена"]
        soon_notifications = [item for item in notifications if item["status"] == "Скоро"]
        archived_notifications = [item for item in notifications if item["status"] == "Архив"]

        if active_filter == "soon":
            filtered_notifications = soon_notifications
        elif active_filter == "archive":
            filtered_notifications = archived_notifications
        else:
            active_filter = "need"
            filtered_notifications = need_notifications

        return render_template(
            "car_notifications.html",
            notifications=filtered_notifications,
            active_filter=active_filter,
            need_replacement_count=len(need_notifications),
            soon_count=len(soon_notifications),
            archive_count=len(archived_notifications),
        )
        
    @app.route("/car/notifications/to-work", methods=["POST"])
    @login_required
    def car_notification_to_work():
        service_name = request.form.get("service_name", "").strip()
        detail_description = request.form.get("detail_description", "").strip()
        period_type = request.form.get("period_type", "").strip()
        work_kind = request.form.get("work_kind", "").strip()
        notification_key = request.form.get("notification_key", "").strip()
        notification_status = request.form.get("notification_status", "").strip()
        last_service_date_text = request.form.get("last_service_date_text", "").strip()

        if not service_name:
            flash("Не удалось перенести уведомление в работу.", "error")
            return redirect(url_for("car_notifications"))

        create_car_planned_service(
            service_name=service_name,
            detail_description=detail_description,
            work_kind=work_kind,
            period_type=period_type,
        )

        if notification_key:
            archive_car_notification(
                notification_key=notification_key,
                title=service_name,
                status=notification_status,
                period_type=period_type,
                detail_description=detail_description,
                last_service_date_text=last_service_date_text,
                work_kind=work_kind,
            )
            hide_car_notification(notification_key)

        flash("Уведомление перенесено в планируемые работы.", "success")
        return redirect(url_for("car_notifications"))

    @app.route("/car/notifications/hide", methods=["POST"])
    @login_required
    def car_notification_hide():
        notification_key = request.form.get("notification_key", "").strip()

        if not notification_key:
            flash("Не удалось скрыть уведомление.", "error")
            return redirect(url_for("car_notifications"))

        hide_car_notification(notification_key)
        flash("Уведомление скрыто.", "success")
        return redirect(url_for("car_notifications"))

    @app.route("/car/notifications/delete-selected", methods=["POST"])
    @login_required
    def car_notifications_delete_selected():
        active_filter = request.form.get("filter", "need").strip() or "need"
        selected_keys = [
            token.strip()
            for token in request.form.get("selected_ids", "").split(",")
            if token.strip()
        ]
        if not selected_keys:
            flash("Выбери хотя бы одно уведомление.", "warning")
            return redirect(url_for("car_notifications", filter=active_filter))

        for notification_key in selected_keys:
            if active_filter == "archive":
                delete_archived_car_notification(notification_key)
            else:
                hide_car_notification(notification_key)

        flash(f"Удалено уведомлений: {len(selected_keys)}.", "success")
        return redirect(url_for("car_notifications", filter=active_filter))

    @app.route("/car/notifications/delete-all", methods=["POST"])
    @login_required
    def car_notifications_delete_all():
        active_filter = request.form.get("filter", "need").strip() or "need"
        notifications = _build_car_notifications()
        if active_filter == "archive":
            filtered = [item for item in notifications if item["status"] == "Архив"]
            if not filtered:
                flash("На этой вкладке нет уведомлений для удаления.", "warning")
                return redirect(url_for("car_notifications", filter=active_filter))
            for item in filtered:
                delete_archived_car_notification(item["notification_key"])
        elif active_filter == "soon":
            filtered = [item for item in notifications if item["status"] == "Скоро"]
            if not filtered:
                flash("На этой вкладке нет уведомлений для удаления.", "warning")
                return redirect(url_for("car_notifications", filter=active_filter))
            for item in filtered:
                hide_car_notification(item["notification_key"])
        else:
            filtered = [item for item in notifications if item["status"] == "Нужна замена"]
            active_filter = "need"
            if not filtered:
                flash("На этой вкладке нет уведомлений для удаления.", "warning")
                return redirect(url_for("car_notifications", filter=active_filter))
            for item in filtered:
                hide_car_notification(item["notification_key"])

        flash("Все уведомления на вкладке удалены.", "success")
        return redirect(url_for("car_notifications", filter=active_filter))
